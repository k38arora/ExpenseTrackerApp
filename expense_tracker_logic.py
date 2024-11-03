# expense_tracker_logic.py

import os
import pandas as pd
import calendar
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.chart import PieChart, Reference
from expense import Expense


class ExpenseTrackerLogic:
    def __init__(self, expense_file_path="expenses.csv"):
        self.expense_file_path = expense_file_path
        self.budget = None
        if self.is_budget_set():
            self.budget = self.get_budget_from_file()

    def set_budget(self, budget):
        """Sets or updates the budget in the expense file."""
        self.budget = budget
        with open(self.expense_file_path, "w") as f:
            f.write(f"Budget, {budget}\n")

    def update_budget(self, new_budget):
        """Updates the budget in the expense file, keeping expenses intact."""
        self.budget = new_budget
        with open(self.expense_file_path, "r") as f:
            lines = f.readlines()[1:]  # Skip the first line (old budget)
        with open(self.expense_file_path, "w") as f:
            f.write(f"Budget, {new_budget}\n")
            f.writelines(lines)

    def is_budget_set(self):
        """Checks if a budget is already set in the expense file."""
        return (
            os.path.exists(self.expense_file_path)
            and self.get_budget_from_file() is not None
        )

    def get_budget_from_file(self):
        """Reads the budget from the CSV file's first line."""
        try:
            with open(self.expense_file_path, "r") as f:
                first_line = f.readline().strip()
                return float(first_line.split(",")[1])
        except (IndexError, ValueError, FileNotFoundError):
            return None

    def add_expense(self, expense):
        """Saves an expense to the file."""
        with open(self.expense_file_path, "a") as f:
            f.write(f"{expense.name}, {expense.amount}, {expense.category}\n")

    def summarize_expenses(self):
        """Returns expenses by category, total spent, and remaining budget."""
        expenses = []
        with open(self.expense_file_path, "r") as f:
            lines = f.readlines()[1:]  # Skip the budget line
            for line in lines:
                name, amount, category = line.strip().split(",")
                expenses.append(Expense(name, category, float(amount)))

        expenses_by_category = {}
        for expense in expenses:
            expenses_by_category[expense.category] = (
                expenses_by_category.get(expense.category, 0) + expense.amount
            )

        total_spent = sum(e.amount for e in expenses)
        remaining_budget = self.budget - total_spent
        return expenses_by_category, total_spent, remaining_budget

    def calculate_daily_budget(self):
        """Calculates the remaining daily budget based on days left in the current month."""
        today = datetime.date.today()
        last_day_of_month = calendar.monthrange(today.year, today.month)[1]
        remaining_days = last_day_of_month - today.day
        _, _, remaining_budget = self.summarize_expenses()

        # Avoid division by zero if there are no remaining days
        daily_budget = (
            remaining_budget / remaining_days
            if remaining_days > 0
            else remaining_budget
        )
        return daily_budget

    def export_to_excel(self, file_path):
        """Exports expenses and summary to an Excel file with a dynamic pie chart."""
        # Ensure expenses exist before exporting
        if not os.path.exists(self.expense_file_path):
            raise FileNotFoundError(
                "No expenses file found. Please add expenses before exporting."
            )

        # Summarize expenses for export
        expenses_by_category, total_spent, remaining_budget = self.summarize_expenses()

        # Prepare detailed expenses data
        expense_details = []
        with open(self.expense_file_path, "r") as f:
            lines = f.readlines()[1:]  # Skip the first line (budget)
            for line in lines:
                name, amount, category = line.strip().split(",")
                expense_details.append([name, category, float(amount)])

        # Create DataFrame for expenses
        expense_df = pd.DataFrame(
            expense_details, columns=["Expense Name", "Category", "Amount"]
        )

        # Export data to Excel
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Write expenses to Excel
            expense_df.to_excel(writer, sheet_name="Expenses", index=False)

            # Write summary to Excel
            worksheet = writer.sheets["Expenses"]
            start_row = len(expense_df) + 2  # Position the summary below the data

            worksheet.cell(row=start_row, column=1).value = "Summary"
            worksheet.cell(row=start_row + 1, column=1).value = (
                f"Total Spent: ${total_spent:.2f}"
            )
            worksheet.cell(row=start_row + 2, column=1).value = (
                f"Remaining Budget: ${remaining_budget:.2f}"
            )

            # Daily budget calculation
            daily_budget = self.calculate_daily_budget()
            worksheet.cell(row=start_row + 3, column=1).value = (
                f"Daily Budget: ${daily_budget:.2f}"
            )

        # Add a dynamic pie chart to Excel
        self.add_pie_chart_to_excel(file_path, expenses_by_category, start_row + 5)

        return file_path

    def add_pie_chart_to_excel(self, file_path, expenses_by_category, start_row):
        """Adds a dynamic pie chart to the Excel file based on expense categories."""
        wb = load_workbook(file_path)
        ws = wb["Expenses"]

        # Write categories and amounts for chart data
        chart_data_row = start_row
        ws.cell(row=chart_data_row, column=1, value="Category")
        ws.cell(row=chart_data_row, column=2, value="Amount")
        for i, (category, amount) in enumerate(
            expenses_by_category.items(), start=chart_data_row + 1
        ):
            ws.cell(row=i, column=1, value=category)
            ws.cell(row=i, column=2, value=amount)

        # Define chart data range
        chart = PieChart()
        labels = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=i)
        data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=i)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Expenses by Category"

        # Position the chart below the summary
        ws.add_chart(chart, f"C{start_row}")

        # Save the workbook with the chart
        wb.save(file_path)
